"""
pact_hud_pic.py — HUD Picture of Subsidized Households occupancy analysis

Reads HUD PIC annual project-level Excel files (2015–2025) and filters to
NYC public housing rows.  Fuzzy-matches each PACT development against the
Public Housing rows to recover pre-conversion occupancy rates and identifies
the non-PACT NYCHA baseline occupancy for every year.

Two goals:
  1. Correct the eviction rate denominator: occupied units rather than total
     contractual units (currently pct_occupied is not used anywhere in the
     pipeline).
  2. Measure vacancy rates: NYCHA-operated vs. PACT-operated over time.

Outputs:
  hud_ph_all.csv        — every NYC Public Housing row from every year
  hud_pact_matches.csv  — fuzzy matches: PACT dev → HUD PH rows, with score
                          and unit-count delta
  hud_summary.json      — annual pct_occupied: non_pact aggregate + per-PACT-
                          dev pre-conversion occupancy

Usage:
  python pact_hud_pic.py [--hud-dir PATH]

Default HUD dir: ~/Desktop/0/W/Work/NYCHA PACT
"""

import csv, json, os, re, sys
from collections import defaultdict

from rapidfuzz import process as rfprocess, fuzz

_HERE   = os.path.dirname(os.path.abspath(__file__))
REF_CSV = os.path.join(_HERE, 'pact_reference.csv')
BBL_CSV = os.path.join(_HERE, 'pact_bbl_master.csv')

DEFAULT_HUD_DIR = os.path.expanduser('~/Desktop/0/W/Work/NYCHA PACT')

OUT_ALL     = os.path.join(_HERE, 'hud_ph_all.csv')
OUT_MATCHES = os.path.join(_HERE, 'hud_pact_matches.csv')
OUT_SUMMARY = os.path.join(_HERE, 'hud_summary.json')

# NYC ZIP prefixes: Manhattan 100-102, Staten Island 103, Bronx 104,
# Queens 110/113-116, Brooklyn 112
NYC_ZIPS = ('100', '101', '102', '103', '104', '110', '112', '113', '114', '115', '116')

FUZZY_THRESHOLD = 72   # token_sort_ratio; lower = more matches, more noise
UNIT_TOL_PCT    = 0.40  # accept if HUD units within ±40% of ref units

# Manual name overrides: PACT ref name → HUD PH name substring to match on.
# Used when fuzzy matching fails due to different naming conventions.
# Each entry is (pact_ref_name, hud_name_fragment) — the fragment is matched
# case-insensitively as a substring against all HUD PH names.
MANUAL_OVERRIDES = {
    'BETANCES I':                      'DR. RAMON E. BETANCES I',
    'BUSHWICK II (GROUPS A & C)':      'HOPE GARDENS',
    'SAMUEL (CITY)':                   'FREDERICK SAMUEL',
    '1010 EAST 178TH STREET':         '1010 E. 178TH',
    'EAGLE AVENUE-EAST 163RD STREET': 'UNION AVE/E 163RD SITE 5',
}


# ── helpers ───────────────────────────────────────────────────────────────────

def _normalize(name):
    """Lowercase, strip punctuation/articles for fuzzy comparison."""
    n = name.lower()
    n = re.sub(r"[^a-z0-9 ]", ' ', n)
    n = re.sub(r'\b(the|a|an|and|of|at|in|for|st|ave|blvd|rd|dr)\b', ' ', n)
    n = re.sub(r'\s+', ' ', n).strip()
    return n


def _year_from_path(path):
    m = re.search(r'PROJECT_(\d{4})', os.path.basename(path))
    return m.group(1) if m else 'unknown'


def _is_nyc_zip(z):
    return str(z or '').startswith(NYC_ZIPS)


# ── loaders ───────────────────────────────────────────────────────────────────

def load_pact_ref():
    """Return list of {dev_name, ref_units, conv_date, status_normalized}."""
    devs = []
    with open(REF_CSV) as f:
        for row in csv.DictReader(f):
            name = row.get('development_name', '').strip()
            if not name:
                continue
            devs.append({
                'dev_name':   name,
                'ref_units':  int(row.get('total_units') or 0),
                'conv_date':  row.get('conversion_date', '').strip(),
                'status':     row.get('status_normalized', '').strip(),
            })
    return devs


def load_bbl_conv_dates():
    """Return {dev_name: conversion_date} from pact_bbl_master.csv (supplements ref)."""
    dates = {}
    with open(BBL_CSV) as f:
        for row in csv.DictReader(f):
            name = row.get('development_name', '').strip()
            d    = row.get('conversion_date', '').strip()
            if name and d:
                dates[name] = d
    return dates


def read_ph_rows_from_file(path):
    """Return list of dicts for every NYC Public Housing row in one Excel file.

    Deduplicates by code — the 2025 source file contains every row twice.
    """
    try:
        import openpyxl
    except ImportError:
        sys.exit('openpyxl is required: pip install openpyxl')

    year = _year_from_path(path)
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb.active
    raw_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx  = {h: i for i, h in enumerate(raw_headers) if h is not None}

    seen_codes = set()
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not str(row[idx.get('states', 0)] or '').startswith('NY'):
            continue
        if not _is_nyc_zip(row[idx.get('STD_ZIP5', 76)]):
            continue
        if row[idx.get('program_label', 5)] != 'Public Housing':
            continue

        code = str(row[idx.get('code', 9)] or '').strip()
        if code in seen_codes:
            continue
        seen_codes.add(code)

        pct_occ = row[idx.get('pct_occupied', 11)]
        # pct_occ of -4 or -5 is HUD's sentinel for suppressed/invalid data
        try:
            pct_occ_val = float(pct_occ) if pct_occ is not None else None
        except (TypeError, ValueError):
            pct_occ_val = None
        if pct_occ_val is not None and pct_occ_val < 0:
            pct_occ_val = None

        rows.append({
            'year':         year,
            'hud_name':     str(row[idx.get('name', 8)] or '').strip(),
            'code':         code,
            'total_units':  int(row[idx.get('total_units', 10)] or 0),
            'pct_occupied': pct_occ_val,
            'zip':          str(row[idx.get('STD_ZIP5', 76)] or '').strip(),
            'address':      str(row[idx.get('STD_ADDR', 74)] or '').strip(),
        })
    wb.close()
    return rows


# ── matching ──────────────────────────────────────────────────────────────────

def fuzzy_match_pact(pact_devs, all_ph_rows):
    """
    For each PACT development, find the best-matching HUD PH row in each year.

    Strategy:
      1. If dev is in MANUAL_OVERRIDES, use substring match against HUD names.
      2. Otherwise fuzzy match by normalized name (token_sort_ratio >= FUZZY_THRESHOLD).
      3. Verify: HUD units within ±UNIT_TOL_PCT of the PACT ref units.
      4. Among remaining candidates pick the highest score.

    Returns list of match dicts.
    """
    # Build {year → list of PH rows}
    by_year = defaultdict(list)
    for r in all_ph_rows:
        by_year[r['year']].append(r)

    # {hud_name_normalized → [ph_rows]} per year for fast lookup
    norm_map_by_year = {}
    for yr, rows in by_year.items():
        nm = defaultdict(list)
        for r in rows:
            nm[_normalize(r['hud_name'])].append(r)
        norm_map_by_year[yr] = nm

    matches = []
    for dev in pact_devs:
        dev_name_norm = _normalize(dev['dev_name'])
        ref_units     = dev['ref_units']
        manual_target = MANUAL_OVERRIDES.get(dev['dev_name'])

        for yr, rows in by_year.items():
            nm = norm_map_by_year[yr]
            candidates = list(nm.keys())
            if not candidates:
                continue

            if manual_target:
                # Substring match against raw HUD names
                target_lower = manual_target.lower()
                ph_rows = [r for r in rows if target_lower in r['hud_name'].lower()]
                if not ph_rows:
                    continue
                score = 100.0
            else:
                result = rfprocess.extractOne(
                    dev_name_norm,
                    candidates,
                    scorer=fuzz.token_sort_ratio,
                    score_cutoff=FUZZY_THRESHOLD,
                )
                if result is None:
                    continue
                best_norm, score, _ = result
                ph_rows = nm[best_norm]

            if ref_units:
                ph_rows_ok = [
                    r for r in ph_rows
                    if abs(r['total_units'] - ref_units) / ref_units <= UNIT_TOL_PCT
                ]
                if not ph_rows_ok:
                    ph_rows_ok = ph_rows
                    unit_ok = False
                else:
                    unit_ok = True
            else:
                ph_rows_ok = ph_rows
                unit_ok = None

            row = max(ph_rows_ok, key=lambda r: r['total_units'])

            matches.append({
                'dev_name':     dev['dev_name'],
                'ref_units':    ref_units,
                'conv_date':    dev['conv_date'],
                'status':       dev['status'],
                'year':         yr,
                'hud_name':     row['hud_name'],
                'hud_code':     row['code'],
                'hud_units':    row['total_units'],
                'pct_occupied': row['pct_occupied'],
                'unit_delta':   row['total_units'] - ref_units if ref_units else None,
                'unit_ok':      unit_ok,
                'match_score':  round(score, 1),
                'zip':          row['zip'],
                'address':      row['address'],
            })

    return matches


# ── summary ───────────────────────────────────────────────────────────────────

def build_summary(all_ph_rows, matches):
    """
    Build hud_summary.json:
      - non_pact_nycha: {year → weighted avg pct_occupied, total_units}
      - pact_pre_conv:  {dev_name → {year → pct_occupied}} (only pre-conversion years)
    """
    # Set of (year, hud_code) pairs that are PACT-matched
    pact_codes_by_year = defaultdict(set)
    for m in matches:
        if m['unit_ok'] is not False:  # exclude unit-mismatch-flagged rows
            pact_codes_by_year[m['year']].add(m['hud_code'])

    # Non-PACT aggregate by year
    non_pact_by_year = defaultdict(lambda: {'weighted_occ': 0.0, 'total_units': 0, 'n_rows': 0})
    for r in all_ph_rows:
        yr   = r['year']
        code = r['code']
        if code in pact_codes_by_year[yr]:
            continue  # exclude PACT-matched rows
        pct = r['pct_occupied']
        if pct is None:
            continue
        units = r['total_units'] or 0
        non_pact_by_year[yr]['weighted_occ'] += pct * units
        non_pact_by_year[yr]['total_units']  += units
        non_pact_by_year[yr]['n_rows']       += 1

    non_pact_out = {}
    for yr in sorted(non_pact_by_year.keys()):
        d = non_pact_by_year[yr]
        tot = d['total_units']
        non_pact_out[yr] = {
            'pct_occupied_wavg': round(d['weighted_occ'] / tot, 1) if tot else None,
            'total_units':       tot,
            'n_developments':    d['n_rows'],
        }

    # Per-PACT-dev, pre-conversion occupancy
    pact_pre = defaultdict(dict)
    for m in matches:
        if m['unit_ok'] is False:
            continue
        conv = m['conv_date']
        yr   = m['year']
        # Only include years that are pre-conversion (or unknown conv date)
        if conv and yr >= conv[:4]:
            continue  # post-conversion year — exclude (data no longer reliable)
        if m['pct_occupied'] is not None:
            pact_pre[m['dev_name']][yr] = {
                'pct_occupied': m['pct_occupied'],
                'hud_units':    m['hud_units'],
                'match_score':  m['match_score'],
            }

    return {
        'non_pact_nycha': non_pact_out,
        'pact_pre_conversion': dict(pact_pre),
    }


# ── main ──────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    hud_dir = DEFAULT_HUD_DIR
    if '--hud-dir' in sys.argv:
        i = sys.argv.index('--hud-dir')
        hud_dir = sys.argv[i + 1]

    # Collect Excel files (deduplicate; prefer no " (1)" duplicate)
    xlsx_files = {}
    for fname in os.listdir(hud_dir):
        if not fname.endswith('.xlsx') or not fname.startswith('PROJECT_'):
            continue
        yr = _year_from_path(fname)
        if yr == 'unknown':
            continue
        # Prefer the clean filename (no " (1)")
        if yr not in xlsx_files or '(1)' not in fname:
            xlsx_files[yr] = os.path.join(hud_dir, fname)

    print(f'Found {len(xlsx_files)} HUD PIC files: {sorted(xlsx_files.keys())}')

    print('\nReading NYC Public Housing rows from each file...')
    all_ph_rows = []
    for yr in sorted(xlsx_files.keys()):
        path = xlsx_files[yr]
        rows = read_ph_rows_from_file(path)
        all_ph_rows.extend(rows)
        units = sum(r['total_units'] for r in rows)
        print(f'  {yr}: {len(rows):3d} rows  {units:,} units')

    print(f'\nTotal PH rows: {len(all_ph_rows):,}')

    with open(OUT_ALL, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=[
            'year', 'hud_name', 'code', 'total_units', 'pct_occupied', 'zip', 'address',
        ])
        w.writeheader()
        w.writerows(all_ph_rows)
    print(f'Wrote {OUT_ALL}')

    print('\nLoading PACT reference...')
    pact_devs    = load_pact_ref()
    bbl_dates    = load_bbl_conv_dates()
    # Supplement ref conv dates from BBL master (which has manual dates for Under Construction)
    for dev in pact_devs:
        if not dev['conv_date'] and dev['dev_name'] in bbl_dates:
            dev['conv_date'] = bbl_dates[dev['dev_name']]
    print(f'  {len(pact_devs)} PACT developments')

    print('\nFuzzy matching PACT devs → HUD PH rows...')
    matches = fuzzy_match_pact(pact_devs, all_ph_rows)
    print(f'  {len(matches):,} match pairs across all years')

    # Print match quality summary
    matched_devs = {m['dev_name'] for m in matches}
    unit_ok_devs = {m['dev_name'] for m in matches if m['unit_ok'] is not False}
    print(f'  {len(matched_devs)} unique PACT devs matched by name')
    print(f'  {len(unit_ok_devs)} pass unit-count verification (±{UNIT_TOL_PCT:.0%})')
    unmatched = [d['dev_name'] for d in pact_devs if d['dev_name'] not in matched_devs]
    if unmatched:
        print(f'  Unmatched: {unmatched}')

    with open(OUT_MATCHES, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=[
            'dev_name', 'ref_units', 'conv_date', 'status',
            'year', 'hud_name', 'hud_code', 'hud_units',
            'pct_occupied', 'unit_delta', 'unit_ok', 'match_score',
            'zip', 'address',
        ])
        w.writeheader()
        w.writerows(sorted(matches, key=lambda m: (m['dev_name'], m['year'])))
    print(f'Wrote {OUT_MATCHES}')

    print('\nBuilding summary...')
    summary = build_summary(all_ph_rows, matches)
    with open(OUT_SUMMARY, 'w') as f:
        json.dump(summary, f, indent=2)
    print(f'Wrote {OUT_SUMMARY}')

    print('\n── Non-PACT NYCHA occupancy by year ──')
    for yr, d in sorted(summary['non_pact_nycha'].items()):
        print(f'  {yr}: {d["pct_occupied_wavg"]}% occupied  '
              f'({d["total_units"]:,}u across {d["n_developments"]} developments)')

    print('\n── PACT pre-conversion occupancy (sample) ──')
    for dev_name, yr_data in sorted(summary['pact_pre_conversion'].items())[:10]:
        yrs = ', '.join(f'{y}:{v["pct_occupied"]}%' for y, v in sorted(yr_data.items()))
        print(f'  {dev_name}: {yrs}')

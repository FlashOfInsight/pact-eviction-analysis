"""
build_hud_nyc.py — Build a clean long-format NYC public housing CSV from HUD PIC

Reads all 11 HUD Picture of Subsidized Households annual Excel files (2015–2025),
filters to NYC public housing rows, deduplicates, and stacks into a single
long-format CSV with one row per development per year.

Output: hud_nyc_public_housing.csv

Usage:
  python build_hud_nyc.py [--hud-dir PATH]

Default HUD dir: ~/Desktop/0/W/Work/NYCHA PACT
"""

import csv, os, re, sys

_HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_HUD_DIR = os.path.expanduser('~/Desktop/0/W/Work/NYCHA PACT')
OUT_CSV = os.path.join(_HERE, 'hud_nyc_public_housing.csv')

# NYC ZIP code prefixes (all 5 boroughs)
# Manhattan: 100-102 | Staten Island: 103 | Bronx: 104
# Queens: 110/113-116 | Brooklyn: 112
NYC_ZIPS = ('100', '101', '102', '103', '104', '110', '112', '113', '114', '115', '116')

# Columns to extract from each file, in output order.
# (output_name, source_column_name_in_excel)
COLUMNS = [
    # Identity
    ('year',            None),          # added programmatically
    ('name',            'name'),
    ('hud_code',        'code'),
    ('pha_entity',      'entities'),
    ('address',         'STD_ADDR'),
    ('city',            'STD_CITY'),
    ('zip',             'STD_ZIP5'),
    ('latitude',        'latitude'),
    ('longitude',       'longitude'),
    # Units & occupancy
    ('total_units',     'total_units'),
    ('pct_occupied',    'pct_occupied'),
    ('people_per_unit', 'people_per_unit'),
    ('people_total',    'people_total'),
    # Reporting quality
    ('pct_reported',    'pct_reported'),
    ('months_since_report', 'months_since_report'),
    # Mobility
    ('pct_movein',      'pct_movein'),       # % who moved in during year
    ('months_from_movein', 'months_from_movein'),  # avg tenure
    ('months_waiting',  'months_waiting'),   # avg waitlist time
    # Income & rent
    ('hh_income',       'hh_income'),        # avg annual household income
    ('rent_per_month',  'rent_per_month'),   # avg rent paid
    ('pct_lt30_median', 'pct_lt30_median'),  # % earning <30% AMI
    # Household composition
    ('pct_female_head',       'pct_female_head'),
    ('pct_female_head_child', 'pct_female_head_child'),
    ('pct_disabled_all',      'pct_disabled_all'),
    ('pct_age62plus',         'pct_age62plus'),
    # Race/ethnicity
    ('pct_minority',          'pct_minority'),
    ('pct_black_nonhsp',      'pct_black_nonhsp'),
    ('pct_hispanic',          'pct_hispanic'),
    # Bedroom mix
    ('pct_overhoused',        'pct_overhoused'),
]

OUTPUT_FIELDS = [col[0] for col in COLUMNS]


def _year_from_path(path):
    m = re.search(r'PROJECT_(\d{4})', os.path.basename(path))
    return m.group(1) if m else 'unknown'


def _clean_numeric(val):
    """Return float or None; treat HUD sentinel negatives as None."""
    if val is None:
        return None
    try:
        f = float(val)
        return None if f < 0 else f
    except (TypeError, ValueError):
        return None


def extract_rows(path):
    try:
        import openpyxl
    except ImportError:
        sys.exit('openpyxl required: pip install openpyxl')

    year = _year_from_path(path)
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb.active
    raw_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx  = {h: i for i, h in enumerate(raw_headers) if h is not None}

    seen_codes = set()
    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        # State filter
        if not str(row[idx.get('states', 0)] or '').startswith('NY'):
            continue
        # NYC ZIP filter
        if not str(row[idx.get('STD_ZIP5', 76)] or '').startswith(NYC_ZIPS):
            continue
        # Public housing only
        if row[idx.get('program_label', 5)] != 'Public Housing':
            continue
        # Deduplicate by code (2025 file has every row doubled)
        code = str(row[idx.get('code', 9)] or '').strip()
        if code in seen_codes:
            continue
        seen_codes.add(code)

        out = {}
        for output_name, src_name in COLUMNS:
            if output_name == 'year':
                out['year'] = year
                continue
            raw = row[idx[src_name]] if src_name in idx else None
            # Numeric fields: clean sentinels
            if src_name in ('total_units', 'pct_occupied', 'people_per_unit',
                            'people_total', 'pct_reported', 'months_since_report',
                            'pct_movein', 'months_from_movein', 'months_waiting',
                            'hh_income', 'rent_per_month', 'pct_lt30_median',
                            'pct_female_head', 'pct_female_head_child',
                            'pct_disabled_all', 'pct_age62plus',
                            'pct_minority', 'pct_black_nonhsp', 'pct_hispanic',
                            'pct_overhoused', 'latitude', 'longitude'):
                out[output_name] = _clean_numeric(raw)
            else:
                out[output_name] = str(raw or '').strip()
        rows.append(out)

    wb.close()
    return rows


if __name__ == '__main__':
    hud_dir = DEFAULT_HUD_DIR
    if '--hud-dir' in sys.argv:
        i = sys.argv.index('--hud-dir')
        hud_dir = sys.argv[i + 1]

    # Collect files, prefer clean filename over " (1)" duplicates
    xlsx_files = {}
    for fname in os.listdir(hud_dir):
        if not fname.endswith('.xlsx') or not fname.startswith('PROJECT_'):
            continue
        yr = _year_from_path(fname)
        if yr == 'unknown':
            continue
        if yr not in xlsx_files or '(1)' not in fname:
            xlsx_files[yr] = os.path.join(hud_dir, fname)

    print(f'Processing {len(xlsx_files)} HUD PIC files...')
    print(f'Output: {OUT_CSV}')
    print()

    all_rows = []
    for yr in sorted(xlsx_files.keys()):
        rows = extract_rows(xlsx_files[yr])
        all_rows.extend(rows)
        occ_vals = [r['pct_occupied'] for r in rows if r['pct_occupied'] is not None]
        wavg = (sum(o * (r['total_units'] or 0) for o, r in zip(occ_vals, [
            r for r in rows if r['pct_occupied'] is not None
        ])) / max(sum(r['total_units'] or 0 for r in rows if r['pct_occupied'] is not None), 1))
        print(f'  {yr}: {len(rows):3d} developments  '
              f'{sum(int(r["total_units"] or 0) for r in rows):,} units  '
              f'avg occupancy {wavg:.1f}%')

    with open(OUT_CSV, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=OUTPUT_FIELDS)
        w.writeheader()
        w.writerows(all_rows)

    print()
    # Panel summary
    codes_by_year = {}
    for r in all_rows:
        codes_by_year.setdefault(r['year'], set()).add(r['hud_code'])
    all_codes = set.union(*codes_by_year.values())
    in_all    = set.intersection(*codes_by_year.values())

    print(f'Total rows:                     {len(all_rows):,}')
    print(f'Unique developments (all years):{len(all_codes):>5}')
    print(f'Present in every year:          {len(in_all):>5}')
    print(f'Appear in some years only:      {len(all_codes) - len(in_all):>5}')
    print(f'  (mostly PACT conversions leaving the public housing program)')
    print()
    print(f'Wrote {OUT_CSV}')
